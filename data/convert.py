#!/usr/bin/env python3
"""
Convert covariance scoring Excel file to whitespace-delimited text files.

Output format (one .txt per sheet):
  Line 1: row count
  Line 2+: mz1  mz2  covariance  partial_cov  score  ranking

Columns 1-2 are formatted as fixed-point decimals (%.8f).
Columns 3-5 are formatted in Fortran-style scientific notation (e.g. 0.56872291E+05).
Column 6 (ranking) is an integer.

The output can be read back with:
    df = pd.read_csv(path, sep=r"\\s+", skiprows=1, header=None, engine="python")
"""

import sys
import os
import openpyxl


def to_fortran_sci(value, decimal_digits=8):
    """
    Format a float in Fortran-style scientific notation: 0.xxxxxxxE+xx
    e.g. 56872.291 -> 0.56872291E+05
         -0.81010740 -> -0.81010740E+00
    """
    if value == 0.0:
        return f"0.{'0' * decimal_digits}E+00"

    sign = "-" if value < 0 else ""
    absval = abs(value)

    # Get the exponent such that 0.1 <= mantissa < 1.0
    # value = mantissa * 10^exponent, where 0.1 <= mantissa < 1.0
    import math
    exp = math.floor(math.log10(absval)) + 1  # shift so mantissa is 0.1..1.0
    mantissa = absval / (10.0 ** exp)

    # Format mantissa with desired digits
    mant_str = f"{mantissa:.{decimal_digits}f}"

    # Format exponent as +XX or -XX
    exp_sign = "+" if exp >= 0 else "-"
    exp_str = f"{abs(exp):02d}"

    return f"{sign}{mant_str}E{exp_sign}{exp_str}"


def convert_sheet(ws, output_path):
    """Convert one worksheet to a whitespace-delimited text file."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header row
        vals = row[:6]  # columns A-F only
        if vals[0] is None:
            continue
        mz1, mz2, cov, pcov, score, ranking = vals
        rows.append((float(mz1), float(mz2), float(cov), float(pcov),
                      float(score), int(ranking)))

    with open(output_path, "w") as f:
        # Line 1: row count (right-aligned to match screenshot indent)
        f.write(f"  {len(rows)}\n")

        for mz1, mz2, cov, pcov, score, ranking in rows:
            col1 = f"{mz1:.8f}"
            col2 = f"{mz2:.8f}"
            col3 = to_fortran_sci(cov)
            col4 = to_fortran_sci(pcov)
            col5 = to_fortran_sci(score)
            col6 = str(ranking)

            # Right-align columns with fixed widths to match screenshot layout
            line = (f"{col1:>16s}  {col2:>16s}  {col3:>20s}  {col4:>20s}  "
                    f"{col5:>20s}  {col6:>8s}")
            f.write(line + "\n")

    print(f"  Written: {output_path}  ({len(rows)} rows)")


def main():
    if len(sys.argv) < 2:
        print("Usage: python convert_excel_to_txt.py <input.xlsx> [output_dir]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(input_path) or "."

    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.load_workbook(input_path, read_only=True)
    print(f"Found {len(wb.sheetnames)} sheets in {input_path}")

    for name in wb.sheetnames:
        ws = wb[name]
        safe_name = name.replace("/", "_").replace("\\", "_")
        output_path = os.path.join(output_dir, f"{safe_name}.txt")
        print(f"  Converting sheet: {name}")
        convert_sheet(ws, output_path)

    wb.close()
    print("Done!")


if __name__ == "__main__":
    main()